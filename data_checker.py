import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import threading
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
import json
import urllib3
import time
import os
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from openpyxl.styles import Alignment


class DataCheckerApp:

    def __init__(self, root):
        self.root = root
        self.root.title("Data Checker")
        self.root.geometry("700x600")

        self.file_path = None
        self.workbook = None
        self.selected_sheets = []

        self.stop_requested = False
        self.start_time = None

        self.build_ui()

    def build_ui(self):
        container = tk.Frame(self.root)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container)
        scrollbar = tk.Scrollbar(container, orient="vertical",
                                command=canvas.yview)

        self.scrollable_frame = tk.Frame(canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0),
                            window=self.scrollable_frame,
                            anchor="nw")

        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        tk.Button(self.scrollable_frame, text="Upload Excel File",
                  command=self.load_file).pack(pady=10)

        self.file_label = tk.Label(self.scrollable_frame,
                                text="Loaded File: None",
                                font=("Arial", 10, "bold"))
        self.file_label.pack(pady=5)

        self.sheet_frame = tk.LabelFrame(self.scrollable_frame, text="Select Sheets")
        self.sheet_frame.pack(fill="both", padx=10, pady=10)

        tk.Label(self.scrollable_frame, text="Starting Cell (e.g., B2)").pack()
        self.start_cell_entry = tk.Entry(self.scrollable_frame)
        self.start_cell_entry.pack()

        tk.Label(self.scrollable_frame, text="Range row to be checked").pack()

        range_frame = tk.Frame(self.scrollable_frame)
        range_frame.pack()

        tk.Label(range_frame, text="From:").grid(row=0, column=0)
        self.range_from_entry = tk.Entry(range_frame, width=10)
        self.range_from_entry.insert(0, "")  # default start
        self.range_from_entry.grid(row=0, column=1)

        tk.Label(range_frame, text="To:").grid(row=0, column=2)
        self.range_to_entry = tk.Entry(range_frame, width=10)
        self.range_to_entry.insert(0, "")  # default end
        self.range_to_entry.grid(row=0, column=3)

        tk.Label(self.scrollable_frame, text="API Endpoint").pack()
        self.api_entry = tk.Entry(self.scrollable_frame, width=80)
        self.api_entry.pack()

        tk.Label(self.scrollable_frame, text="Bearer Token").pack()
        self.token_entry = tk.Entry(self.scrollable_frame, width=80)
        self.token_entry.pack()

        self.progress = ttk.Progressbar(self.scrollable_frame, length=500)
        self.progress.pack(pady=10)

        self.time_label = tk.Label(self.scrollable_frame, text="Time: 00:00")
        self.time_label.pack(pady=5)

        tk.Label(self.scrollable_frame, text="Process Log").pack()

        self.log_text = tk.Text(self.scrollable_frame, height=10, bg="black", fg="lime")
        self.log_text.pack(fill="both", padx=10, pady=5)

        self.log_text.configure(state="disabled")

        tk.Label(self.scrollable_frame, text="Workers:").pack()

        self.worker_entry = tk.Entry(self.scrollable_frame, width=10)
        self.worker_entry.insert(0, "20")   # default 20 workers
        self.worker_entry.pack()

        self.start_button = tk.Button(
            self.scrollable_frame,
            text="Start Checking",
            command=self.start_checking
        )
        self.start_button.pack(pady=10)

        self.stop_button = tk.Button(
            self.scrollable_frame,
            text="Stop",
            command=self.stop_process,
            state="disabled"
        )
        self.stop_button.pack(pady=5)

        self.reset_button = tk.Button(
            self.scrollable_frame,
            text="Reset",
            command=self.reset_app
        )
        self.reset_button.pack(pady=5)

    def load_file(self):
        self.file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")]
        )

        if not self.file_path:
            return

        # Show loading
        self.file_label.config(text="Loading Excel file...")
        self.root.update_idletasks()

        def load_workbook_thread():
            try:
                self.workbook = load_workbook(self.file_path)

                self.root.after(0, self.populate_sheets)
                self.root.after(0, lambda:
                    self.file_label.config(
                        text=f"Loaded File: {os.path.basename(self.file_path)}"
                    )
                )
                filename = os.path.basename(self.file_path)
                self.file_label.config(text=f"Loaded File: {filename}")

            except Exception as e:
                self.root.after(0, lambda:
                    messagebox.showerror("Error", str(e))
                )

        threading.Thread(target=load_workbook_thread).start()

    def populate_sheets(self):
        for widget in self.sheet_frame.winfo_children():
            widget.destroy()

        self.sheet_vars = {}

        for sheet in self.workbook.sheetnames:
            var = tk.BooleanVar()
            chk = tk.Checkbutton(self.sheet_frame, text=sheet, variable=var)
            chk.pack(anchor="w")
            self.sheet_vars[sheet] = var

    def reset_app(self):
        self.progress["value"] = 0
        self.log_text.delete(1.0, tk.END)

        self.stop_requested = False
        self.start_time = None

        self.file_label.config(text="No file selected")

    def stop_process(self):
        self.stop_requested = True
        self.log("Stop requested by user...")

    def update_timer(self):
        if self.start_time and not self.stop_requested:
            elapsed = int(time.time() - self.start_time)
            mins = elapsed // 60
            secs = elapsed % 60
            self.time_label.config(text=f"Time: {mins:02d}:{secs:02d}")
            self.root.after(1000, self.update_timer)

    def log(self, message):
        self.root.after(0, self._append_log, message)

    def _append_log(self, message):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

        with open("debug.log", "a", encoding="utf-8") as f:
            f.write(message + "\n")

    def start_checking(self):
        self.selected_sheets = [
            sheet for sheet, var in self.sheet_vars.items() if var.get()
        ]

        if not self.selected_sheets:
            messagebox.showerror("Error", "Please select at least one sheet.")
            return

        confirm = messagebox.askyesno(
            "Confirm",
            f"Sheets: {self.selected_sheets}\n"
            f"API: {self.api_entry.get()}\n\nProceed?"
        )

        if not confirm:
            return

        self.stop_requested = False
        self.start_time = time.time()

        self.update_timer()

        self.start_button.config(state="disabled")
        self.stop_button.config(state="normal")

        thread = threading.Thread(target=self.process_data)
        thread.start()

    def bulk_check_assets(self, start_cell, api_url, token, limit=None):
        session = requests.Session()
        session.headers.update({
            "Authorization": f"Bearer {token}"
        })

        green_fill = PatternFill(start_color="00FF00",
                                end_color="00FF00",
                                fill_type="solid")

        yellow_fill = PatternFill(start_color="FFFF00",
                                end_color="FFFF00",
                                fill_type="solid")

        red_fill = PatternFill(start_color="FF9999",
                            end_color="FF9999",
                            fill_type="solid")

        col = ''.join(filter(str.isalpha, start_cell))
        row_start = int(''.join(filter(str.isdigit, start_cell)))

        tasks = []

        for sheet in self.selected_sheets:
            ws = self.workbook[sheet]

            range_from_val = self.range_from_entry.get().strip()
            range_to_val = self.range_to_entry.get().strip()

            if range_from_val:
                row_from = int(range_from_val)
            else:
                row_from = row_start

            if range_to_val:
                row_to = int(range_to_val)
            else:
                row_to = ws.max_row

            for row in range(row_from, row_to + 1):
                cell_value = ws[f"{col}{row}"].value
                if not cell_value:
                    continue

                asset_code = str(cell_value).strip()
                tasks.append((sheet, row, asset_code))

        total = len(tasks)

        if total == 0:
            raise Exception("No data found in selected range.")

        # Data collectors for report
        match_data = []
        mismatch_data = []
        not_found_data = []

        def check_one(task):
            if self.stop_requested:
                return None
            
            sheet, row, asset_code = task
            ws = self.workbook[sheet]

            excel_name = (ws.cell(row=row, column=6).value or "").strip()
            excel_category = (ws.cell(row=row, column=7).value or "").strip()

            params = {
                "page": 1,
                "limit": 10,
                "search": asset_code
            }

            try:
                full_url = requests.Request(
                    "GET",
                    api_url,
                    params=params
                ).prepare().url

                self.log("=" * 80)
                self.log(f"REQUESTING:")
                self.log(full_url)

                r = session.get(
                    api_url,
                    params=params,
                    timeout=5,
                    verify=False
                )

                self.log(f"STATUS CODE: {r.status_code}")

                if r.status_code == 200:
                    json_data = r.json()

                    json_response_str = json.dumps(json_data, indent=2)
                    self.log("SERVER RESPONSE:")
                    self.log(json_response_str)

                    if json_data.get("success"):
                        asset_list = json_data.get("data", {}).get("data", [])
                        self.log(f"ASSET LIST LENGTH: {len(asset_list)}")

                        for item in asset_list:
                            if item.get("code") == asset_code:

                                api_name = (item.get("name") or "").strip()
                                api_category = (item.get("assetCategoryTera") or "").strip()

                                self.log(f"Excel Name: {excel_name}")
                                self.log(f"API Name: {api_name}")
                                self.log(f"Excel Category: {excel_category}")
                                self.log(f"API Category: {api_category}")

                                if api_name == excel_name and api_category == excel_category:
                                    return (sheet, row, "match", asset_code,
                                            excel_name, excel_category,
                                            api_name, api_category, json_response_str)

                                return (sheet, row, "mismatch", asset_code,
                                        excel_name, excel_category,
                                        api_name, api_category, json_response_str)

            except Exception as e:
                self.log("EXCEPTION OCCURRED:")
                self.log(str(e))
                return (sheet, row, "error", asset_code,
                        excel_name, excel_category,
                        None, None)
            
            self.log(f"{asset_code} -> NOT FOUND")
            return (sheet, row, "not_found", asset_code,
                    excel_name, excel_category,
                    None, None, "NOT FOUND IN SYSTEM")

        processed = 0

        try:
            workers = int(self.worker_entry.get())
        except:
            workers = 20

        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = []

            # Submit tasks gradually (important for cancellation)
            for task in tasks:
                if self.stop_requested:
                    self.log("Process stopped before submitting all tasks.")
                    break
                futures.append(executor.submit(check_one, task))

            for future in as_completed(futures):

                if self.stop_requested:
                    self.log("Process stopping... waiting running tasks to finish.")
                    break

                result = future.result()

                # If worker exited early due to stop
                if result is None:
                    continue

                (sheet, row, status, asset_code,
                ex_name, ex_cat, api_name,
                api_cat, json_response) = result

                ws = self.workbook[sheet]

                if status == "match":
                    fill_style = green_fill
                    match_data.append(
                        (row, asset_code, ex_name, ex_cat, json_response)
                    )

                elif status == "mismatch":
                    fill_style = yellow_fill
                    mismatch_data.append(
                        (row, asset_code, ex_name, ex_cat, json_response)
                    )

                elif status == "not_found":
                    fill_style = red_fill
                    not_found_data.append(
                        (row, asset_code, ex_name, ex_cat, json_response)
                    )

                else:
                    fill_style = red_fill

                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=c).fill = fill_style

                processed += 1

                self.root.after(
                    0,
                    lambda p=processed: self.progress.configure(
                        value=(p / total) * 100
                    )
                )

            # Optional: cancel remaining futures that haven't started
            if self.stop_requested:
                for f in futures:
                    f.cancel()

        return total, match_data, mismatch_data, not_found_data

    def generate_report(self, total, match_data,
                    mismatch_data, not_found_data):
        report_wb = Workbook()

        report_wb.remove(report_wb.active)

        summary = report_wb.create_sheet("SUMMARY")

        match_count = len(match_data)
        mismatch_count = len(mismatch_data)
        not_found_count = len(not_found_data)

        success_rate = (match_count / total) * 100 if total else 0

        summary["A1"] = f"Total Checked : {total}"
        summary["A2"] = f"Exact Match   : {match_count}"
        summary["A3"] = f"Mismatch      : {mismatch_count}"
        summary["A4"] = f"Not Found     : {not_found_count}"
        summary["A5"] = f"Success Rate  : {success_rate:.2f}%"

        nf_sheet = report_wb.create_sheet("NOT_FOUND")
        nf_sheet["A1"] = f"Total Not Found: {not_found_count}"
        nf_sheet.append(["Row", "Asset Code", "Name", "Category", "JSON Response"])

        for row in not_found_data:
            nf_sheet.append(row)

        mm_sheet = report_wb.create_sheet("MISMATCH")
        mm_sheet["A1"] = f"Total Mismatch: {mismatch_count}"
        mm_sheet.append(["Row", "Asset Code", "Name", "Category", "JSON Response"])

        for row in mismatch_data:
            mm_sheet.append(row)

        match_sheet = report_wb.create_sheet("MATCH")
        match_sheet["A1"] = f"Total Exact Match: {match_count}"
        match_sheet.append(["Row", "Asset Code", "Name", "Category", "JSON Response"])

        for row in match_data:
            match_sheet.append(row)

        base_name = "Report_" + os.path.basename(self.file_path)
        name, ext = os.path.splitext(base_name)

        output_report = base_name
        counter = 2

        while os.path.exists(output_report):
            output_report = f"{name} ({counter}){ext}"
            counter += 1

        for sheet in report_wb.worksheets:
            sheet.column_dimensions["A"].width = 8     # Row
            sheet.column_dimensions["B"].width = 22    # Asset Code
            sheet.column_dimensions["C"].width = 30    # Name
            sheet.column_dimensions["D"].width = 25    # Category
            sheet.column_dimensions["E"].width = 80    # JSON

            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

        report_wb.save(output_report)

        return output_report

    def process_data(self):
        start_cell = self.start_cell_entry.get()
        api_url = self.api_entry.get()
        token = self.token_entry.get()

        try:
            total, match_data, mismatch_data, not_found_data = \
                self.bulk_check_assets(start_cell, api_url, token)

            report_file = self.generate_report(
                total, match_data, mismatch_data, not_found_data
            )

            self.root.after(
                0,
                lambda: messagebox.showinfo("Done", f"Report saved as:\n{report_file}")
            )
            self.root.after(0, lambda: self.start_button.config(state="normal"))
            self.root.after(0, lambda: self.stop_button.config(state="disabled"))
        except Exception as e:
            self.root.after(
                0,
                lambda: messagebox.showerror("Error", str(e))
            )

if __name__ == "__main__":
    root = tk.Tk()
    app = DataCheckerApp(root)
    root.mainloop()